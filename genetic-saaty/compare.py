import numpy as np
from openpyxl import load_workbook

BASE_PATH = "d:/research/code/genetic-saaty/"

ORIGINAL_EXCEL = BASE_PATH + "filtered_matrices_CR_010_080.xlsx"
REDUCED_EXCEL = BASE_PATH + "ga_reduced_n8_matrices.xlsx"
OUTPUT_EXCEL = BASE_PATH + "ga_reduced_n8_with_measures.xlsx"

N = 8
TOL = 1e-6

RI = {
    1: 0.00, 2: 0.00, 3: 0.58, 4: 0.90,
    5: 1.12, 6: 1.24, 7: 1.32, 8: 1.41,
    9: 1.45, 10: 1.49
}


def ahp_metrics(A):
    eigvals, eigvecs = np.linalg.eig(A)
    idx = np.argmax(eigvals.real)

    lambda_max = eigvals[idx].real

    w = eigvecs[:, idx].real
    w = np.abs(w)
    w = w / np.sum(w)

    CI = (lambda_max - A.shape[0]) / (A.shape[0] - 1)
    CR = CI / RI[A.shape[0]]

    return lambda_max, CI, CR, w


def get_headers(ws):
    return {cell.value: cell.column for cell in ws[1]}


def read_matrix_from_row(ws, row, headers, n=8):
    values = []

    for i in range(1, n + 1):
        for j in range(1, n + 1):
            col_name = f"A{i}{j}"
            values.append(ws.cell(row=row, column=headers[col_name]).value)

    return np.array(values, dtype=float).reshape(n, n)


def kendall_tau_distance(w1, w2):
    r1 = list(np.argsort(-w1))
    r2 = list(np.argsort(-w2))

    pos2 = {item: idx for idx, item in enumerate(r2)}

    tau = 0
    for i in range(len(r1)):
        for j in range(i + 1, len(r1)):
            a = r1[i]
            b = r1[j]

            if pos2[a] > pos2[b]:
                tau += 1

    return tau


def preference_measures(A_original, A_reduced):
    _, _, CR_original, w_original = ahp_metrics(A_original)
    _, _, CR_reduced, w_reduced = ahp_metrics(A_reduced)

    d = np.mean(np.abs(w_original - w_reduced))
    d_percent = d * 100

    D = np.sum(np.abs(A_original - A_reduced))

    N_changed = np.sum(np.abs(A_original - A_reduced) > TOL)

    tau = kendall_tau_distance(w_original, w_reduced)

    return d, d_percent, D, N_changed, tau, CR_original, CR_reduced


def main():
    wb_original = load_workbook(ORIGINAL_EXCEL, data_only=True)
    ws_original = wb_original["n8_filtered"]

    wb_reduced = load_workbook(REDUCED_EXCEL)
    ws_reduced = wb_reduced.active

    headers_original = get_headers(ws_original)
    headers_reduced = get_headers(ws_reduced)

    new_columns = [
        "CR_before",
        "CR_after",
        "d",
        "d_percent",
        "D",
        "N_changed",
        "Kendall_tau"
    ]

    start_col = ws_reduced.max_column + 1

    for c, name in enumerate(new_columns, start=start_col):
        ws_reduced.cell(row=1, column=c).value = name

    max_rows = min(ws_original.max_row, ws_reduced.max_row)

    for row in range(2, max_rows + 1):
        A_original = read_matrix_from_row(ws_original, row, headers_original, N)
        A_reduced = read_matrix_from_row(ws_reduced, row, headers_reduced, N)

        d, d_percent, D, N_changed, tau, CR_before, CR_after = preference_measures(
            A_original,
            A_reduced
        )

        values = [
            CR_before,
            CR_after,
            d,
            d_percent,
            D,
            N_changed,
            tau
        ]

        for c, value in enumerate(values, start=start_col):
            ws_reduced.cell(row=row, column=c).value = value

    wb_reduced.save(OUTPUT_EXCEL)

    print("Done.")
    print("Saved:", OUTPUT_EXCEL)


if __name__ == "__main__":
    main()